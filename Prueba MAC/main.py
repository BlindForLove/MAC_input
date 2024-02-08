import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QMessageBox
from mi_interfaz import Ui_MainWindow
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re


class MiVentana(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.pushButton.clicked.connect(self.introducir_valores)

        # Obtener el número de serie actual del archivo de Excel
        self.contador_serie = self.obtener_numero_serie()

        # Establecer el número de serie inicial en el cuadro de texto
        self.ui.textEdit.setText(str(self.contador_serie))

    def introducir_valores(self):
        # Obtener los valores de los cuadros de texto
        numero_serie = self.ui.textEdit.toPlainText()
        direccion_mac = self.ui.textEdit_2.toPlainText()

        if not self.validar_direccion_mac(direccion_mac):
            QMessageBox.warning(self, 'Error de validación', 'Por favor, introduzca una dirección MAC válida en formato "XX:XX:XX:XX:XX:XX".')
            return

        # Crear un archivo Excel o cargar uno existente
        try:
            workbook = load_workbook('MAC SONDAS.xlsx')
        except FileNotFoundError:
            workbook = Workbook()

        # Obtener la hoja activa o crear una nueva
        sheet = workbook.active if workbook.sheetnames else workbook.create_sheet()

        # Obtener la próxima fila disponible
        #next_column = len(sheet[1]) + 1
        next_row = self.obtener_proxima_fila_disponible(sheet, column=1)

        # Poner Encabezados
        #if next_row == 1:
        #    sheet.cell(row=next_row, column=1).value = "Nº SERIE"
        #    sheet.cell(row=next_row, column=16).value = "DIRECCIÓN MAC"

        # Escribir los valores en las celdas correspondientes
        sheet.cell(row=next_row, column=1).value = int(numero_serie)
        sheet.cell(row=next_row, column=16).value = direccion_mac

        # Verificar si es la primera fila y escribir los encabezados
        #if next_column == 1:
        #    sheet.cell(row=1, column=1).value = "Nº SERIE"
        #    sheet.cell(row=1, column=16).value = "DIRECCIÓN MAC"
            # Establecer el formato de la celda como número para el número de serie
         #   sheet.column_dimensions[get_column_letter(next_column)].number_format = '0'

        # Escribir los valores en las celdas correspondientes
        #sheet.cell(row=2, column=1).value = int(numero_serie)
        #sheet.cell(row=2, column=16).value = direccion_mac

        workbook.save('MAC SONDAS.xlsx')
        print("Valores guardados en MAC SONDAS.xlsx")

        # Incrementar el contador de serie
        self.contador_serie += 1

        # Actualizar el cuadro de texto del número de serie
        self.ui.textEdit.setText(str(self.contador_serie))

         # Borrar el campo "Dirección MAC"
        self.ui.textEdit_2.clear()

    def validar_direccion_mac(self, direccion_mac):
        regex = re.compile(r'^([0-9A-Fa-f]{2}:){5}[0-9A-Fa-f]{2}$')
        return bool(regex.match(direccion_mac))
    
    def obtener_numero_serie(self):
        try:
            workbook = load_workbook('MAC SONDAS.xlsx')
        except FileNotFoundError:
            return 100

        sheet = workbook.active if workbook.sheetnames else None

        if sheet:
            rows = sheet.iter_rows(values_only=True)
            next(rows)  # Omitir la primera fila (encabezados)
            last_row = max((row[0] for row in rows if row[0] and isinstance(row[0], int)), default=0)
          


            return int(last_row) + 1
        else:
            return 100
    def obtener_proxima_fila_disponible(self, sheet, column):
        for row in sheet.iter_rows(min_row=1, min_col=column, max_col=column):
            if row[0].value is None:
                return row[0].row
        return sheet.max_row + 1

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ventana = MiVentana()
    ventana.show()
    sys.exit(app.exec_())
